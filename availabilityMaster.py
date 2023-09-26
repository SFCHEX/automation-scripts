import pandas as pd
import openpyxl
import argparse 
from datetime import datetime , timedelta
from openpyxl import load_workbook
import os

avaSheetMasterName="C:\\Users\\swx1283483\\Desktop\\tes\\Sep-Network Availability Dashboard-2023-18 (F).xlsx"
excludedSitesSheet="C:\\Users\\swx1283483\\Desktop\\tes\\Excluded List From Cells Breakdown.xlsx"


outputName="excludedAvailibilityOutsideCairo.xlsx"
sheetName='OC Cells AVA (All Tech)'



def get_files_in_current_directory():
    try:
        current_directory = os.getcwd()
        files = os.listdir(current_directory)
        return [x for x in files if x.endswith(".xlsx") and outputName != x]
    except Exception as e:
        print(f"Error getting files in current directory: {str(e)}")
        return []

def merge_files(input_files):
    merged_data = pd.DataFrame()
    for input_file in input_files:
        df = pd.read_excel(input_file,"OC Cells AVA (All Tech)")
        merged_data =pd.concat([merged_data,df])
    return merged_data


def excludeAVA(avaSheet):
    headers=["DAYID","CELL","REG","SITE",'TECH',"VENDOR","DOWN_TIME","AVA (%)"]
    avaSheet=avaSheet[avaSheet["TECH"]!="4G"]
    avaSheet=avaSheet[headers]
    excludedSheet= pd.read_excel(excludedSitesSheet, sheet_name="Sheet1")
    ZTEUpdateSheet= pd.read_excel(excludedSitesSheet, sheet_name="ZTE UPDATE")
    excluded=excludedSheet["SITE"].tolist()
    ZTEUpdate=ZTEUpdateSheet["Site ID"].tolist()
    avaSheet=avaSheet[~avaSheet["SITE"].isin(excluded)]
    avaSheetZTE=avaSheet[avaSheet["VENDOR"]=="Z"]
    avaSheetREST=avaSheet[avaSheet["VENDOR"]!="Z"]
    avaSheetZTE=avaSheetZTE[avaSheet["SITE"].isin(ZTEUpdate)]
    avaSheet=pd.concat([avaSheetZTE,avaSheetREST])
    return avaSheet


def analysis(avaSheet):
    
    avaSheet["D-Loss"]= 100-avaSheet["AVA (%)"]
    avaSheetAvgAVA=avaSheet["AVA (%)"].mean()

    avgAVAPivot = pd.pivot_table(avaSheet, values=r'AVA (%)', index='REG',aggfunc='mean')
    new_row=pd.Series({"AVA (%)":avaSheetAvgAVA},name="Total")
    avgAVAPivot=avgAVAPivot.append(new_row,ignore_index=False)
    avgAVAPivot["Loss"]=100-avgAVAPivot["AVA (%)"]
    avgAVAPivot["Loss+avgAVA"]=avgAVAPivot["AVA (%)"]+avgAVAPivot["Loss"]

    avgLoss=avgAVAPivot["Loss"].mean()

    avgAVA=avgAVAPivot["AVA (%)"].mean()
    avg=avgAVAPivot["Loss+avgAVA"].mean()
    totalAVG=pd.DataFrame({"":["Total Average"],"AVA (%)" :[avgAVA], "Loss":[avgLoss],"Loss+avgAVA":[avg] })
#    avgAVAPivot=avgAVAPivot.append(new_row,ignore_index=True)

    avaLossSum=avaSheet["D-Loss"].sum()

    avaSheet["D-Loss*W"]=avaSheet["D-Loss"].apply(lambda x: (x/avaLossSum) * avgLoss )
    
    finalWeightPivot = pd.pivot_table(avaSheet, values=r'D-Loss*W', index='SITE',aggfunc={"D-Loss*W":['sum',"count"]})
    finalWeightPivot.reset_index(inplace=True)
    avgAVAPivot.reset_index(inplace=True)

    return avaSheet,avgAVAPivot,finalWeightPivot,totalAVG

def updateNetworkAva(avaSheetName):
    try:
        avaWBMaster=openpyxl.load_workbook(avaSheetMasterName)
        sheetnum=2
        avaSheetMaster=avaWBMaster[f'OC Cells AVA (All Tech)-{sheetnum}']
        avaWBSheet=openpyxl.load_workbook(avaSheetName)
        avaSheet=avaWBSheet['Sheet1']
        print("opened AVA master sheet")
        for row in avaSheet.iter_rows(min_row=2, values_only=True):
                avaSheetMaster.append(row)
        avaWBMaster.save(avaSheetMasterName)
        print("Added values to master sheet")
    except Exception as e:
        print(e)

def saveSheet(avaSheet,avgAVA,finalWeightPivot,totalAVG):
    with pd.ExcelWriter(outputName,engine="xlsxwriter") as writer:
        avaSheet.to_excel(writer,sheet_name=sheetName,index=False)
        avgAVA.to_excel(writer,sheet_name="Analysis",index=False)
        finalWeightPivot.to_excel(writer,sheet_name="Analysis",index=False,startcol=5)
        totalAVG.to_excel(writer,sheet_name="Analysis",index=False,startrow=7)


def main():
    parser = argparse.ArgumentParser(description="DataFornetAv")
    parser.add_argument("-a", "--avaSheetName", help="availability Update")
    args =parser.parse_args()

    if args.avaSheetName:
        print("Running AVAILABILITYMASTER")
        avaSheet=excludeAVA(args.avaSheetName)
        avaSheet,avgAVA,finalWeightPivot,totalAVG=analysis(avaSheet)
        saveSheet(avaSheet,avgAVA,finalWeightPivot,totalAVG)

        print("outputed CellBreakdownExcluded.xlsx")
    else:
        print("Running AVAILABILITYMASTER")
        input_files=get_files_in_current_directory()
        avaSheet=merge_files(input_files)
        avaSheet=excludeAVA(avaSheet)
        avaSheet,avgAVA,finalWeightPivot,totalAVG=analysis(avaSheet)
        saveSheet(avaSheet,avgAVA,finalWeightPivot,totalAVG)

        print("outputed CellBreakdownExcluded.xlsx")

 



if __name__ == "__main__":

    main()
