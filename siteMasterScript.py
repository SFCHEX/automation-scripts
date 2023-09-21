import pandas as pd
import openpyxl
import argparse 
from datetime import datetime , timedelta
from openpyxl import load_workbook


avaSheetMasterName="C:\\Users\\swx1283483\\Desktop\\tes\\Sep-Network Availability Dashboard-2023-18 (F).xlsx"
powerSheetMasterName="C:\\Users\\swx1283483\\Desktop\\tes\\Sep_2023 Daily Commercial Power alarm Outside-18 (F).xlsx"
outputSheetName="C:\\Users\\swx1283483\\Desktop\\tes\\Sep-Sites AVA,PAR &amp; PAD.xlsx"
excludedSitesSheet="C:\\Users\\swx1283483\\Desktop\\tes\\Excluded List From Cells Breakdown.xlsx"


# Define the vendor mapping
vendor_mapping = {
    "HUAWEI": "HUAWEI",
    "NOKIA": "NOKIA",
    "ZTE": "ZTE"
}

def findVendor(file_name):
    for vendor, keyword in vendor_mapping.items():
        if keyword.lower() in file_name.lower():
            return vendor

def mergePower(input_files,output_file):
    merged_data = pd.DataFrame()
    for input_file in input_files:
        df = pd.read_excel(input_file)
        vendor = findVendor(input_file)
        df["Vendor"] = vendor 
        df["Date"] = datetime.now().date()  
        df["MTTR.1"] =df['MTTR'].apply(lambda x:x.hour + x.minute /  60 + x.second /  3600)
        merged_data =pd.concat([merged_data,df])

    merged_data["ALARM"]=merged_data["Alarm Name"]
    merged_data = merged_data[['Region', 'Site Name', 'Tier', 'Site Code',"ID","Alarm Name","Occurrence Time(NT)","Clearance Time(NT)","MTTR","Type","Date", 'MTTR.1',"Vendor","ALARM"]]

    AlarmSheet= pd.read_excel(powerSheetMasterName, sheet_name="Sheet2")
    ALARM=[item.lower() for item in AlarmSheet["Alarm Name"].tolist()]
    print(len(ALARM))
    merged_data["ALARM"]=merged_data["ALARM"].apply(lambda x:"Yes" if x.lower() in ALARM else "No")


    
    merged_data.to_excel(output_file, index=False)
    print(f"Data merged and saved to {output_file}")

def loadSheets(avaSheetName,powerSheetName):
    avaSheet=pd.read_excel(avaSheetName,"OC Cells AVA (All Tech)")

    excludedSheet= pd.read_excel(excludedSitesSheet, sheet_name="Sheet1")
    ZTEUpdateSheet= pd.read_excel(excludedSitesSheet, sheet_name="ZTE UPDATE")

   
    excluded=excludedSheet["SITE"].tolist()
    ZTEUpdate=ZTEUpdateSheet["Site ID"].tolist()
    
    avaSheet=avaSheet[~avaSheet["SITE"].isin(excluded)]

    avaSheetZTE=avaSheet[avaSheet["VENDOR"]=="Z"]
    avaSheetREST=avaSheet[avaSheet["VENDOR"]!="Z"]
    avaSheetZTE=avaSheetZTE[avaSheet["SITE"].isin(ZTEUpdate)]
    avaSheet=pd.concat([avaSheetZTE,avaSheetREST])

    avaSheet.to_excel("CellBreakdownExcluded.xlsx", sheet_name='OC Cells AVA (All Tech)', index=False)
    powerSheet=pd.read_excel(powerSheetName,"Sheet1")
    return avaSheet, powerSheet

def updateNetworkAva(avaSheetName):
    try:
        avaWBMaster=openpyxl.load_workbook(avaSheetMasterName)
        sheetnum=2
        avaSheetMaster=avaWBMaster[f'OC Cells AVA (All Tech)-{sheetnum}']
        avaWBSheet=openpyxl.load_workbook(avaSheetName)
        avaSheet=avaWBSheet['Sheet1']
        print("opened AVA master sheet")
        for row in avaSheet.iter_rows(min_row=2, values_only=True):
                avaSheetMaster.append(row)
        avaWBMaster.save(avaSheetMasterName)
        print("Added values to master sheet")
    except Exception as e:
        print(e)



def updatePowerAlarm(powerSheetName):
    try:
        powerWBMaster=openpyxl.load_workbook(powerSheetMasterName)
        powerSheetMaster=powerWBMaster['Sheet1']
        powerWBSheet=openpyxl.load_workbook(powerSheetName)
        powerSheet=powerWBSheet['Sheet1']
        print("opened power master sheet")
        for row in powerSheet.iter_rows(min_row=2, values_only=True):
                powerSheetMaster.append(row)
        powerWBMaster.save(powerSheetMasterName)
        print("added values to master sheet")
    except Exception as e:
        print(e)



def getAVA(avaSheet):
    avaSheet=avaSheet.loc[avaSheet["TECH"]!="4G"]
    print(avaSheet)
    AVA = pd.pivot_table(avaSheet, values=r'AVA (%)', index='SITE',aggfunc='mean')
#   blank_columns = pd.DataFrame('', index=range(len(avaSheet)), columns=[''] * 3)  # Insert 3 blank columns
#   avaSheet=pd.concat([avaSheet,blank_columns,avaSheet],axis=1)
    return AVA


def getPADPAR(powerSheet):
    powerSheet=powerSheet.loc[powerSheet['ALARM']=="Yes"]
    print(powerSheet)
    PAD = pd.pivot_table(powerSheet, values='MTTR.1', index='Site Code', aggfunc='sum')
    PAR = pd.pivot_table(powerSheet, values='MTTR.1', index='Site Code', aggfunc='count')
#   blank_columns = pd.DataFrame('', index=range(len(powerSheet)), columns=[''] * 3)  # Insert 3 blank columns
#   powerSheet=pd.concat([powerSheet,blank_columns,PAD,blank_columns,PAR],axis=1)
 
    return PAD,PAR


def updateSiteData(AVA,PAD,PAR,date,output):
    wb=openpyxl.load_workbook(output)
    sheet=wb.active
    print(AVA)
    print(PAD)
    print(PAR)
    dictAVA=AVA.to_dict(orient='index')
    dictPAD=PAD.to_dict(orient='index')
    dictPAR=PAR.to_dict(orient='index')
    data=[dictAVA,dictPAD,dictPAR]
    count=-1

    for row in sheet.iter_rows(min_row=2, max_row=2):
        print(count)
        for cell in row:
            print(f"{str(cell.value)==str(date)}")
            if str(cell.value)==str(date):
                count+=1
                if count==0:
                    func="AVA (%)"
                else:
                    func="MTTR.1"
                print("Modifying sheet now")
                for row_index in range(3, sheet.max_row+1):
                    cell2=sheet[f'A{row_index}']
                    print(f"looping through column {cell.col_idx}")
                    cell2_value=cell2.value
                    changedCell=sheet.cell(column=cell.col_idx, row=row_index)
                    if cell2_value in data[count]:
                        print("changing values")
                        changedCell.value=data[count][cell2_value][func]
                    else:
                        changedCell.value=0
    wb.save(output)


def main():
    parser = argparse.ArgumentParser(description="DataFornetAv")
    parser.add_argument("-a", "--avaSheetName", help="availability Update")
    parser.add_argument("-p", "--powerSheetNames", nargs="+", help="power update")
    parser.add_argument("-t", "--time",help="time")
    args =parser.parse_args()
    mergedPowerSheetName="CombinedPowerCommercialSheet.xlsx"
    mergePower(args.powerSheetNames,mergedPowerSheetName)  
    avaSheet,powerSheet=loadSheets(args.avaSheetName,mergedPowerSheetName)
    
    print("Loaded Sheets")
#    updateNetworkAva(args.avaSheetName)
#    print("Updated Sheet Network")
#    updatePowerAlarm(mergedPowerSheetName)
    print("Updated Sheet power")
    AVA=getAVA(avaSheet)
    print("received AVA")
    PAD,PAR=getPADPAR(powerSheet)
    print("received PADPAR")
    if args.time:
        date=args.time
    else:
        date = (datetime.now()-timedelta(days=1)).strftime("%Y-%m-%d 00:00:00")
    updateSiteData(AVA,PAD,PAR,date,outputSheetName)
    print("Updated Site Data")



if __name__ == "__main__":

    main()
