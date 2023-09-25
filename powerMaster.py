import pandas as pd
import openpyxl
import argparse 
from datetime import datetime , timedelta
from openpyxl import load_workbook
import os

powerSheetMasterName="C:\\Users\\swx1283483\\Desktop\\tes\\Sep_2023 Daily Commercial Power alarm Outside-18 (F).xlsx"


outputName="combinedPowerSheet.xlsx"

vendor_mapping = {
    "HUAWEI": "HUAWEI",
    "NOKIA": "NOKIA",
    "ZTE": "ZTE"
}

def findVendor(file_name):
    for vendor, keyword in vendor_mapping.items():
        if keyword.lower() in file_name.lower():
            return vendor

def get_files_in_current_directory():
    try:
        current_directory = os.getcwd()
        files = os.listdir(current_directory)
        return [x for x in files if x.endswith(".xlsx") and outputName != x]
    except Exception as e:
        print(f"Error getting files in current directory: {str(e)}")
        return []


def mergePower(input_files,output_file):
    merged_data = pd.DataFrame()
    for input_file in input_files:
        df = pd.read_excel(input_file)
        vendor = findVendor(input_file)
        df["Vendor"] = vendor 
        df["Date"] = datetime.now().date()  
        df["MTTR.1"] =df['MTTR'].apply(lambda x:x.hour + x.minute /  60 + x.second /  3600)
        merged_data =pd.concat([merged_data,df])
    merged_data["ALARM"]=merged_data["Alarm Name"]
    merged_data = merged_data[['Region', 'Site Name', 'Tier', 'Site Code',"ID","Alarm Name","Occurrence Time(NT)","Clearance Time(NT)","MTTR","Type","Date", 'MTTR.1',"Vendor","ALARM"]]

    AlarmSheet= pd.read_excel(powerSheetMasterName, sheet_name="Sheet2")
    ALARM=[item.lower() for item in AlarmSheet["Alarm Name"].tolist()]
    merged_data["ALARM"]=merged_data["ALARM"].apply(lambda x:"Yes" if x.lower() in ALARM else "No")


    
    merged_data.to_excel(output_file, index=False)
    print(f"Data merged and saved to {output_file}")


def updatePowerAlarm(powerSheetName):
    try:
        powerWBMaster=openpyxl.load_workbook(powerSheetMasterName)
        powerSheetMaster=powerWBMaster['Sheet1']
        powerWBSheet=openpyxl.load_workbook(powerSheetName)
        powerSheet=powerWBSheet['Sheet1']
        print("opened power master sheet")
        for row in powerSheet.iter_rows(min_row=2, values_only=True):
                powerSheetMaster.append(row)
        powerWBMaster.save(powerSheetMasterName)
        print("added values to master sheet")
    except Exception as e:
        print(e)



def main():
    parser = argparse.ArgumentParser(description="DataFornetAv")
    parser.add_argument("-p", "--powerSheetName", help="power Sheet Name Update")
    args =parser.parse_args()

    if args.powerSheetName:
        mergePower(args.powerSheetName)

    else:
        input_files=get_files_in_current_directory()
        mergePower(input_files,outputName)
   

 



if __name__ == "__main__":

    main()

