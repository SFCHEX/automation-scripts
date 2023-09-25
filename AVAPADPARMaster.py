import pandas as pd
import openpyxl
import argparse 
from datetime import datetime , timedelta
from openpyxl import load_workbook


outputSheetName="C:\\Users\\swx1283483\\Desktop\\Sep-Sites AVA,PAR & PAD.xlsx"


mergedPowerSheetName="combinedPowerSheet.xlsx"
outputExcludedName="excludedAvailibilityOutsideCairo.xlsx"

def loadSheets(avaSheetName,powerSheetName):
    avaSheet=pd.read_excel(avaSheetName,"OC Cells AVA (All Tech)")
    powerSheet=pd.read_excel(powerSheetName,"Sheet1")
    return avaSheet, powerSheet

def getAVA(avaSheet):
    print(avaSheet)
    AVA = pd.pivot_table(avaSheet, values=r'AVA (%)', index='SITE',aggfunc='mean')
#   blank_columns = pd.DataFrame('', index=range(len(avaSheet)), columns=[''] * 3)  # Insert 3 blank columns
#   avaSheet=pd.concat([avaSheet,blank_columns,avaSheet],axis=1)
    return AVA


def getPADPAR(powerSheet):
    powerSheet=powerSheet.loc[powerSheet['ALARM']=="Yes"]
    print(powerSheet)
    PAD = pd.pivot_table(powerSheet, values='MTTR.1', index='Site Code', aggfunc='sum')
    PAR = pd.pivot_table(powerSheet, values='MTTR.1', index='Site Code', aggfunc='count')
#   blank_columns = pd.DataFrame('', index=range(len(powerSheet)), columns=[''] * 3)  # Insert 3 blank columns
#   powerSheet=pd.concat([powerSheet,blank_columns,PAD,blank_columns,PAR],axis=1)
    return PAD,PAR


def updateSiteData(AVA,PAD,PAR,date,output):
    wb=openpyxl.load_workbook(output)
    sheet=wb.active
    print(AVA)
    print(PAD)
    print(PAR)
    dictAVA=AVA.to_dict(orient='index')
    dictPAD=PAD.to_dict(orient='index')
    dictPAR=PAR.to_dict(orient='index')
    data=[dictAVA,dictPAD,dictPAR]
    count=-1

    for row in sheet.iter_rows(min_row=2, max_row=2):
        print(count)
        for cell in row:
            print(f"{str(cell.value)==str(date)}")
            if str(cell.value)==str(date):
                count+=1
                if count==0:
                    func="AVA (%)"
                else:
                    func="MTTR.1"
                print("Modifying sheet now")
                for row_index in range(3, sheet.max_row+1):
                    cell2=sheet[f'A{row_index}']
                    print(f"looping through column {cell.col_idx}")
                    cell2_value=cell2.value
                    changedCell=sheet.cell(column=cell.col_idx, row=row_index)
                    if cell2_value in data[count]:
                        print("changing values")
                        changedCell.value=data[count][cell2_value][func]
                    else:
                        changedCell.value=0
    wb.save(output)


def main():
    parser = argparse.ArgumentParser(description="DataFornetAv")
    parser.add_argument("-a", "--avaSheetName", help="availability Update")
    parser.add_argument("-p", "--powerSheetName", help="power update")
    parser.add_argument("-t", "--time",help="time")
    args =parser.parse_args()
    
    if args.avaSheetName and args.powerSheetName:
        avaSheet,powerSheet=loadSheets(args.avaSheetName,args.powerSheetName)
    else:
        avaSheet,powerSheet=loadSheets(outputExcludedName,mergedPowerSheetName)
 
   
    AVA=getAVA(avaSheet)
    print("received AVA")
    PAD,PAR=getPADPAR(powerSheet)
    print("received PADPAR")
    if args.time:
        date=args.time
    else:
        date = (datetime.now()-timedelta(days=1)).strftime("%Y-%m-%d 00:00:00")
    updateSiteData(AVA,PAD,PAR,date,outputSheetName)
    print("Updated Site Data")




if __name__ == "__main__":

    main()

